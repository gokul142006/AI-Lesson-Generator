"""
excel_export.py
----------------
Generates a professionally formatted Excel workbook (.xlsx) for a
lesson plan, using OpenPyXL. Returns raw bytes so the Streamlit layer
can feed them straight into st.download_button.
"""

from __future__ import annotations

import io
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils import SECTION_SCHEMA, markdown_to_plain_text


class ExportError(Exception):
    """Raised when export generation fails, with a friendly message."""

    def __init__(self, user_message: str):
        super().__init__(user_message)
        self.user_message = user_message


HEADER_FILL = PatternFill(start_color="1F3B57", end_color="1F3B57", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FONT = Font(bold=True, size=16, color="1F3B57")
LABEL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(horizontal="center", vertical="center")


def build_excel_workbook(lesson: Dict[str, Any], meta: Dict[str, Any]) -> bytes:
    """
    Build the Excel workbook in memory and return its bytes.

    meta is expected to contain: subject, grade, topic, duration.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Lesson Plan"

        # --- Title row -----------------------------------------------------
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value = lesson.get("lesson_title") or "Lesson Plan"
        title_cell.font = TITLE_FONT
        title_cell.alignment = CENTER
        ws.row_dimensions[1].height = 30

        # --- Header row for the table --------------------------------------
        headers = ["Section", "Content"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        row = 3

        # --- Metadata rows ---------------------------------------------------
        metadata_rows = [
            ("Subject", meta.get("subject", "")),
            ("Grade", meta.get("grade", "")),
            ("Topic", meta.get("topic", "")),
            ("Duration", meta.get("duration", "")),
        ]
        for label, value in metadata_rows:
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            content_cell = ws.cell(row=row, column=2, value=value)
            content_cell.alignment = WRAP_TOP
            content_cell.border = THIN_BORDER
            row += 1

        # --- Lesson sections --------------------------------------------------
        for key, label in SECTION_SCHEMA:
            if key == "lesson_title":
                continue  # already shown as the sheet title
            content = markdown_to_plain_text(lesson.get(key, ""))

            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = LABEL_FONT
            label_cell.alignment = WRAP_TOP
            label_cell.border = THIN_BORDER

            content_cell = ws.cell(row=row, column=2, value=content)
            content_cell.alignment = WRAP_TOP
            content_cell.border = THIN_BORDER

            # Rough auto-height: taller rows for longer content.
            estimated_lines = max(1, content.count("\n") + len(content) // 60)
            ws.row_dimensions[row].height = min(15 * max(estimated_lines, 2), 400)

            row += 1

        # --- Column sizing -----------------------------------------------------
        ws.column_dimensions[get_column_letter(1)].width = 30
        ws.column_dimensions[get_column_letter(2)].width = 90

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    except Exception as exc:
        raise ExportError(f"Could not generate the Excel file: {exc}")
